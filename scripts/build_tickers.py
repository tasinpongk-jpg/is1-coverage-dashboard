"""
Rebuild data/tickers.json from the team portfolio Excel.

Usage:
  python build_tickers.py "<path to IS1 Port Summary.xlsx>"

Excel must have a single sheet (Sheet1) with columns: Company, Sector, RM Name.
"""

from __future__ import annotations
import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "data" / "tickers.json"

# Map source sector → canonical bucket name (PF&REIT → PFREIT, else identity)
SECTOR_BUCKET = {
    "AGRI": "AGRI", "FOOD": "FOOD", "CONS": "CONS",
    "CONMAT": "CONMAT", "PROP": "PROP", "PF&REIT": "PFREIT",
}


def main():
    if len(sys.argv) < 2:
        sys.exit("Usage: python build_tickers.py <portfolio.xlsx>")
    src = Path(sys.argv[1])
    if not src.exists():
        sys.exit(f"File not found: {src}")

    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]

    tickers = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        tk, sector, rm = row[0], row[1], row[2]
        # Anonymise RM to its initial (privacy): "Champ"->"C", "Orn"->"O", ...
        rm = (str(rm).strip()[:1].upper() if rm not in (None, "") else rm)
        tickers.append({
            "tk": tk,
            "sector": sector,
            "rm": rm,
            "bucket": SECTOR_BUCKET.get(sector, sector),
        })

    rms = sorted({t["rm"] for t in tickers},
                 key=lambda r: -sum(1 for t in tickers if t["rm"] == r))
    sectors = sorted({t["sector"] for t in tickers})

    payload = {
        "version": 1,
        "tickers": tickers,
        "rms": rms,
        "sectors": sectors,
        "totals": {
            "all": len(tickers),
            "by_rm": {r: sum(1 for t in tickers if t["rm"] == r) for r in rms},
            "by_sector": {s: sum(1 for t in tickers if t["sector"] == s) for s in sectors},
        },
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {len(tickers)} tickers across {len(rms)} RMs to {OUT}")
    print(f"By RM:     {payload['totals']['by_rm']}")
    print(f"By sector: {payload['totals']['by_sector']}")


if __name__ == "__main__":
    main()
